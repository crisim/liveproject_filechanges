import os
import ntpath
import sys
import sqlite3
import time
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

"""Ritorna il percorso del database"""
def getbasefile():
    return os.path.splitext(os.path.basename(__file__))[0] 

"""Si connette al db SQLite"""
def connectdb():
    try:
        dbfile = getbasefile() + '.db'
        conn = sqlite3.connect(dbfile, timeout=2)
        return conn
    except sqlite3.OperationalError as err:
        print(str(err))

"""Funzione che esegue query senza connessione alla base"""
def runcmdnoconn(query, args=None):
    result = False
    try:
        conn = connectdb()
        if not conn is None:
            result = runcmd(conn,query,args)
        
    except sqlite3.OperationalError as err:
         print(str(err))
    finally:
        if conn != None:
            conn.close()

    return result

"""Funzione base che esegue query senza risultato"""
def runcmd(conn, query, args=None):
    args = args or []
    result = False
    try:
        cursor = conn.execute(query, args)
        conn.commit()
        cursor.close()
        result = True
    except sqlite3.OperationalError as err:
        print(str(err))

    return result

"""Esiste la tabella?"""
def tableexists(table):
    result = False
    try:
        conn = connectdb()
        if not conn is None:
            query = "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?"
            args = (table,)
            cur = conn.execute(query, args)
            item = cur.fetchone()
            result = not item is None
            cur.close()
            conn.close()
    except:
        print ("tableexist ERROR")
    finally:
        if conn != None:
            conn.close()
    return result

"""Crea tabella"""
def createhashtable():
    if tableexists('files'):
        return True
    query = "CREATE TABLE files (id INTEGER PRIMARY KEY, fname VARCHAR(255) NOT NULL, md5 BLOB NOT NULL)"
    return runcmdnoconn(query)

"""Crea indice"""
def createhashtableidx():
    if tableexists('files'):
        query = 'CREATE INDEX IF NOT EXISTS idxfile ON files (fname)'
        return runcmdnoconn(query)
    return False

"""Update the SQLite File Table"""
def updatehashtable(fname, md5):
    qry = "UPDATE files SET md5 = ? WHERE fname = ?"
    return runcmdnoconn(qry,(md5,fname))

"""Insert into the SQLite File Table"""
def inserthashtable(fname, md5):
    qry = "INSERT INTO files (fname, md5) VALUES (?,?)" 
    return runcmdnoconn(qry, (fname, md5))

"""Setup's the Hash Table"""
def setuphashtable():
   createhashtable()
   createhashtableidx()

"""Checks if md5 hash tag exists in the SQLite DB"""
def md5indb(fname):
    items = []
    if tableexists('files'):
        try:
            conn = connectdb()
            if not conn is None:
                qry = "SELECT md5 FROM files WHERE fname = ?"
                try:
                    cursor = conn.execute(qry,(fname,))
                    for item in cursor:
                        items.append(item[0])
                finally:
                    if cursor != None:
                        cursor.close()
        except sqlite3.OperationalError as ex:
            print(str(ex))
        finally:
            if conn != None:
                conn.close()
    return items

def haschanged(fname, md5):
    items = md5indb(fname)
    if len(items) == 0:
        inserthashtable(fname,md5)
    else:
        if not md5 in items:
            updatehashtable(fname, md5)
            return True

def getfileext(fname):
    """Get the file name extension"""
    return os.path.splitext(fname)[1]

def getmoddate(fname):
    """Get file modified date"""
    mtime = os.path.getmtime(fname)
    return datetime.fromtimestamp(mtime)

def md5short(fname):
    """Get md5 file hash tag"""
    with open(fname, mode='rb') as openfile:
        content = openfile.read()
        result = hashlib.md5(content).digest()
    return result

def loadflds():
    flds = []
    ext = []
    inifile = getbasefile() + '.ini'
    if os.path.isfile(inifile):
        cfile = open(inifile, 'r')
        for line in cfile:
            tokens = line.split("|")
            fld = tokens[0]
            fext = []
            if len(tokens) == 2:
                fext = tokens[1].split(",")
            flds.append(fld)
            ext.append(fext)
    return flds, ext

def checkfilechanges(folder, exclude, ws):
    changed = False
    """Checks for files changes"""
    for subdir, dirs, files in os.walk(folder):
        for fname in files:
            origin = os.path.join(subdir, fname)
            if os.path.isfile(origin):
                if not os.path.splitext(origin)[1] in exclude:
                    md5 = md5short(origin)
                    if haschanged(origin, md5):
                        changed=True
                        fn, fld = ntpath.split(origin)
                        mt = getmoddate(origin)
                        rowxlsreport(ws,fn,origin,fld,mt.strftime("%d-%b-%Y"), mt.strftime("%I:%M:%S"))
                        #write in file
    return changed

def path_leaf(path):
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)

def runfilechanges(ws):
    changed = False
    fldexts = loadflds()
    for i, fld in enumerate(fldexts[0]):
        if checkfilechanges(fld, fldexts[1][i], ws):
            changed=True
    return changed

def execute(args):
    # Start the creation of the Excel report
    wb, ws, st = startxlsreport()
    if '--loop' in args:    
        try:
            while True:
                changed = runfilechanges(ws)
        except KeyboardInterrupt:
            # Check for a keyboard interruption to stop the script
            pass
    else:
        changed = runfilechanges(ws)
    # Finalize the creation of the Excel report
    endxlsreport(wb, st)


def getdt(fmt):
    """Get the current DateTime as a string"""
    now = datetime.now()
    return now.strftime(fmt)


def startxlsreport():
    # Create the workbook, get the hostname and current DateTime
    wb = Workbook()
    ws = wb.active
    st = getdt("%d-%b-%Y %H_%M_%S")
    headerxlsreport(ws)

    return wb, ws, st

def headerxlsreport(ws):
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Full File Name")
    ws.cell(row=1, column=3, value="Folder Name")
    ws.cell(row=1, column=4, value="Date")
    ws.cell(row=1, column=5, value="Time")

    ft = Font(bold=True)
    ws["A1"].font = ft
    ws["A2"].font = ft
    ws["A3"].font = ft
    ws["A4"].font = ft
    ws["A6"].font = ft

def getlastrow(ws):
    return ws.max_row + 1

def rowxlsreport(ws, fn, ffn, fld, d, t):
    row = getlastrow(ws)
    ws.cell(row=row, column=1, value=fn)
    ws.cell(row=row, column=2, value=ffn)
    ws.cell(row=row, column=3, value=fld)
    ws.cell(row=row, column=4, value=d)
    ws.cell(row=row, column=5, value=t)

def endxlsreport(wb, st):
    dt = ' from ' + st + ' to ' + getdt("%d-%b-%Y %H_%M_%S")
    # Finalize the creation of the Excel report
    wb.save(getbasefile() + dt + '.xlsx')

"""Test"""
if __name__ == '__main__':
    execute(sys.argv)
